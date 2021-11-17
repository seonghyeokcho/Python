from openpyxl import load_workbook
wb = load_workbook("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 (국어) 영어 수학
# ws.move_range('B1:C11', rows=0, cols=1)
# ws['B2'].value = '국어'  # B1 cell 에 '국어' 입력

ws.move_range('C1:C11', rows=5, cols=-1)

wb.save("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample_move.xlsx")