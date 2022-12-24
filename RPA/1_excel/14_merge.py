from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells('B2:D2')  # B2 부터 D2 까지 병합
ws['B2'].value = 'Merged Cell'


wb.save("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample_merge.xlsx")