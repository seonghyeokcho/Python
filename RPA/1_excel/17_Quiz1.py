from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string
import time
import collections as coll
#### 두 번째 방법이 효율적

## 첫 번째 방법(나)
start = time.time()
wb = Workbook()
ws = wb.active

# 데이터를 입력(리스트,튜플)
ws.append(['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'])
source = [
    [1,10,8,5,14,26,12],
    [2,7,3,7,15,24,18],
    [3,9,5,8,8,12,4],
    [4,7,8,7,17,21,18],
    [5,7,8,7,16,25,15],
    [6,3,5,8,8,17,0],
    [7,4,9,10,16,27,18],
    [8,6,6,6,15,19,17],
    [9,10,10,9,19,30,19],
    [10,9,8,8,20,25,20]
    ]

# for i in range(10):  # 범위(range)로 지정 - 두번 호출(비효율)
#     ws.append(source[i])
# ↓ 수정
for i in source:  # 리스트를 지정- 한번 호출(효율)
    ws.append(i)

# 1. 퀴즈2 점수를 10으로 수정
# 2중 반복문(for 문) 사용, 두 번째 방법 보다 소모 시간이 작음
for col in ws.iter_cols(min_row=2, min_col=4, max_col=4):
    # print(col)
    for row in col:
        row.value = 10
        # print(row.value)

# 2. H열에 총점(SUM 이용)
# ws.insert_cols(8, 2)  # insert 는 셀 사이에 또 다른 셀을 추가할 때만 사용
ws['H1'] = '총점'
ws['I1'] = '성적'

for row in ws.iter_rows(min_row=2, min_col=2):
    # print(row)
    # print(row[0].coordinate, row[5].coordinate)
    ws[row[6].coordinate] = '=SUM({}:{})'.format(row[0].coordinate, row[5].coordinate)
    # print(ws[row[6].coordinate].value)

sum_score = []
for score in source:
    sum_score.append(sum(score[1:]) - score[3] + 10)
# print(sum_score)

# I열에 성적 정보 추가
# 출석이 5 미만인 학생은 총점 상관없이 F
count = 0
for row in ws.iter_rows(min_row=2, min_col=2):
    # print(ws[row[6].coordinate].value)

    # - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
    if int(ws[row[0].coordinate].value) < 5:  # 출석 5 미만
        ws[row[7].coordinate] = 'F'
    else:
        if sum_score[count] >= 90:  # 총점 90점 이상
            ws[row[7].coordinate] = 'A'
        elif sum_score[count] >= 80:  # 총점 80점 이상
            ws[row[7].coordinate] = 'B'
        elif sum_score[count] >= 70:  # 총점 70점 이상
            ws[row[7].coordinate] = 'C'
        else:
            ws[row[7].coordinate] = 'D'
    count += 1
    # print(ws[row[7].coordinate].value)

print('time :', time.time() - start)

wb.save("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/data/scores.xlsx")

## 두 번째 방법(유튜버) - win
# start = time.time()
# wb = Workbook()
# ws = wb.active

# ws.append(['학번', '출석', '퀴즈1', '퀴즈2', '중간고사', '기말고사', '프로젝트'])
# source = [
#     [1,10,8,5,14,26,12],
#     [2,7,3,7,15,24,18],
#     [3,9,5,8,8,12,4],
#     [4,7,8,7,17,21,18],
#     [5,7,8,7,16,25,15],
#     [6,3,5,8,8,17,0],
#     [7,4,9,10,16,27,18],
#     [8,6,6,6,15,19,17],
#     [9,10,10,9,19,30,19],
#     [10,9,8,8,20,25,20]
#     ]


# for i in source:  # 리스트를 지정- 한번 호출(효율)
#     ws.append(i)

# # 1. 퀴즈2 점수를 10으로 수정
# # 반복문, 조건문을 한번씩 사용, 모든 셀의 데이터를 한번에 교체함, 첫 번째 방법 보다 소모 시간이 큼
# for idx, cell in enumerate(ws['D']):
#     if idx == 0:
#         continue
#     cell.value = 10

# ws['H1'] = '총점'
# ws['I1'] = '성적'

# # 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# for idx, score in enumerate(source, start=2):
#     sum_val = sum(score[1:]) - score[3] + 10  # 총점
#     ws.cell(row=idx, column=8).value = '=SUM(B{}:G{})'.format(idx, idx)

#     # - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D
#     grade = None
#     if sum_val >= 90:
#         grade = 'A'
#     elif sum_val >= 80:
#         grade = 'B'
#     elif sum_val >= 70:
#         grade = 'C'
#     else:
#         grade = 'D'
    
#     # 3. 출석이 5 미만인 학생은 총점 상관없이 F
#     if score[1] < 5:
#         grade = 'F'
    
#     ws.cell(row=idx, column=9).value = grade

# print('time :', time.time() - start)

# wb.save("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/data/scores.xlsx")