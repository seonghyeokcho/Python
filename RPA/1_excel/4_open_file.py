from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample.xlsx")
# sample.xlsx 파일에서 wb를 불러옴
ws = wb.active  # 활성화된 Sheet

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
    #     print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 . . .
    # print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")  # 1 2 3 4 . . .
    print()

# print(ws.max_row, ws.max_column)