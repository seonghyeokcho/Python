from openpyxl import load_workbook
# wb = load_workbook("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 자겨오고 있음
# for row in ws.values:
#     for cell in row:
#         print(cell)

wb = load_workbook("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/data/sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가져옴
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시됨
for row in ws.values:
    for cell in row:
        print(cell)