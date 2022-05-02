from openpyxl import load_workbook
wb = load_workbook("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample_merge.xlsx")
ws = wb.active

# B2:D2 병합되어 있던 셀을 헤제
ws.unmerge_cells('B2:D2')

wb.save("/Users/csh/jupytercreation/Python_Practice/RPA/basic/1_excel/sample_unmerge.xlsx")